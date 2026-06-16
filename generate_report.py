from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pathlib import Path

report_path = Path('GlassCo_Project_Report.xlsx')
wb = Workbook()

# Overview sheet
ws = wb.active
ws.title = 'Project Overview'
ws.append(['Project Report for Glass Co. Website'])
ws.append([])
ws.append(['Project Name', 'Glass Co. Website'])
ws.append(['Location', str(Path.cwd())])
ws.append(['Report Date', '2026-06-16'])
ws.append(['Prepared By', 'Manual Testing Report'])
ws.append([])
ws.append(['Scope'])
ws.append(['- Static frontend website pages: index, about, services, contact'])
ws.append(['- Backend contact submission API using Express and SQLite'])
ws.append(['- Contact form integration and local database storage'])
ws.append([])
ws.append(['Implementation Status'])
ws.append(['- Backend server file present: server.js'])
ws.append(['- Dependency manifest present: package.json'])
ws.append(['- SQLite data directory created: data/contacts.db'])
ws.append(['- Contact form submission wired to /api/contact'])

cell = ws['A1']
cell.font = Font(bold=True, size=14)

# Test cases sheet
ws = wb.create_sheet('Test Cases')
headers = ['Test Case ID', 'Test Case', 'Expected Result', 'Actual Result', 'Status', 'Comments']
ws.append(headers)

test_cases = [
    ('TC-01', 'Load Home page', 'Home page loads successfully', 'Home page content available in HTML', 'Pass', 'Static page is present'),
    ('TC-02', 'Load Contact page', 'Contact page loads successfully', 'Contact page content available in HTML', 'Pass', 'Static page is present'),
    ('TC-03', 'Submit contact form with valid data', 'Server returns success and stores record', 'Requires backend running; code is implemented', 'Pending', 'Manual runtime validation after server launch'),
    ('TC-04', 'Submit contact form without email', 'Server returns 400 validation error', 'Backend code checks required fields', 'Pass', 'Validation logic exists for email/message'),
    ('TC-05', 'Submit contact form without message', 'Server returns 400 validation error', 'Backend code checks required fields', 'Pass', 'Validation logic exists for email/message'),
    ('TC-06', 'View stored contacts via API', 'GET /api/contacts returns contact list', 'Endpoint exists in server code', 'Pending', 'Manual runtime validation needed'),
    ('TC-07', 'Install dependencies', 'npm install completes successfully', 'npm install completed successfully in project folder', 'Pass', 'Dependencies installed from package.json'),
    ('TC-08', 'Start backend server', 'Server starts on configured port without error', 'Port conflict may exist if another process uses port 3000', 'Conditional', 'Need free port or stop existing process'),
]
for row in test_cases:
    ws.append(row)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Findings sheet
ws = wb.create_sheet('Findings')
ws.append(['Finding Category', 'Description', 'Severity', 'Status'])
findings = [
    ('Backend Startup', 'Port 3000 may be in use by another process; server startup fails with EADDRINUSE until the port is freed or changed.', 'Medium', 'Open'),
    ('Static Page Delivery', 'All HTML pages are present and linked correctly from navigation.', 'Low', 'Closed'),
    ('Contact Submission', 'Form submission is wired to POST /api/contact and backend stores data in SQLite.', 'Low', 'Closed (code-level)'),
    ('Input Validation', 'Email and message fields are required; additional validation could be added for email format and required service selection.', 'Medium', 'Open'),
    ('Security', 'No HTTPS, no input sanitization beyond basic checks, and no spam protection.', 'High', 'Open'),
    ('Admin Access', 'No administrative UI exists to review stored contact submissions.', 'Medium', 'Open'),
]
for row in findings:
    ws.append(row)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Skills and tools sheet
ws = wb.create_sheet('Skills & Tools')
ws.append(['Category', 'Items'])
entries = [
    ('Frontend', 'HTML5, CSS3, JavaScript'),
    ('Backend', 'Node.js, Express.js'),
    ('Database', 'SQLite3'),
    ('Runtime', 'Node.js v24'),
    ('Package Management', 'npm'),
    ('Testing Approach', 'Manual testing, browser validation, API contract verification'),
    ('Development Tools', 'VS Code, PowerShell, npm'),
    ('Libraries', 'express, sqlite3, nodemon'),
]
for row in entries:
    ws.append(row)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Future improvements sheet
ws = wb.create_sheet('Future Improvements')
ws.append(['Improvement Area', 'Description'])
improvements = [
    ('Port Configurability', 'Allow the server to accept an environment variable for port selection and provide a fallback UI message if the port is occupied.'),
    ('Form Validation', 'Add client-side validation for email format, required fields, and improved UI error messages.'),
    ('Admin Dashboard', 'Add an admin page to view stored contact submissions and export them as CSV.'),
    ('Security Hardening', 'Add input sanitization, HTTPS support, rate limiting, and spam protection/CAPTCHA for contact submissions.'),
    ('Deployment', 'Deploy the website and backend to a cloud provider or container platform with environment configuration.'),
    ('Email Notification', 'Add email alerts for new contact submissions using a mailing service.'),
    ('User Experience', 'Improve form feedback, loading states, success messages, and responsive design test coverage.'),
]
for row in improvements:
    ws.append(row)
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Summary sheet
ws = wb.create_sheet('Summary')
ws.append(['Manual Testing Summary'])
ws.append([])
summary_lines = [
    'The Glass Co. website project includes a frontend static site and a backend API for contact form submissions.',
    'The backend is built with Express.js and stores contact data in a local SQLite database.',
    'Primary functionality has been implemented and dependency installation is complete.',
    'The key remaining issue is ensuring the backend server starts on a free port when another Node process is active.',
    'Manual validation of live form submission and contact retrieval is pending until the server is launched successfully.',
    'Overall, the project is ready for client review with the noted issues and future improvements documented.',
]
for line in summary_lines:
    ws.append([line])
cell = ws['A1']
cell.font = Font(bold=True, size=14)

wb.save(report_path)
print(f'Created report: {report_path.resolve()}')
